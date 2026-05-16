#!/usr/bin/env python3
"""
nfa_extract.py — Extract noise figure & gain traces from an Agilent 8973A
                  via GPIB (USB-to-IEEE488 adapter) and save to Excel.

Requirements:
    pip install pyvisa openpyxl

VISA backend:
    - NI-VISA  (recommended on Windows, install from ni.com)
    - Keysight IO Libraries
    - pyvisa-py  (pure Python fallback: pip install pyvisa-py)

Usage:
    python nfa_extract.py                        # single capture
    python nfa_extract.py --label "LNA Rev-A"    # label for this DUT
    python nfa_extract.py --append results.xlsx --label "LNA Rev-B"
                                                  # append to existing workbook

The 8973A default GPIB address is 8.  Override with --addr.
"""

import argparse
import sys
import time
import os
from datetime import datetime

import pyvisa
import numpy as np

# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------
def parse_args():
    p = argparse.ArgumentParser(
        description="Extract NF & Gain from Agilent 8973A via GPIB")
    p.add_argument("--addr", type=int, default=8,
                   help="GPIB address of the 8973A (default 8)")
    p.add_argument("--board", type=int, default=0,
                   help="GPIB board index (default 0)")
    p.add_argument("--label", type=str, default=None,
                   help="DUT label for column headers (e.g. 'LNA Rev-A')")
    p.add_argument("--append", type=str, default=None,
                   help="Append results to an existing Excel file")
    p.add_argument("--output", type=str, default=None,
                   help="Output Excel filename (default: nfa_YYYYMMDD_HHMMSS.xlsx)")
    p.add_argument("--timeout", type=int, default=30000,
                   help="VISA timeout in ms (default 30000)")
    p.add_argument("--trigger", action="store_true",
                   help="Trigger a fresh single sweep before reading")
    p.add_argument("--backend", type=str, default="",
                   help="PyVISA backend, e.g. '@py' for pyvisa-py")
    p.add_argument("--list", action="store_true",
                   help="List VISA resources and exit")
    return p.parse_args()


# ---------------------------------------------------------------------------
# GPIB helpers
# ---------------------------------------------------------------------------
def connect(board: int, addr: int, timeout: int, backend: str):
    """Open a VISA session to the 8973A."""
    rm = pyvisa.ResourceManager(backend)

    resource = f"GPIB{board}::{addr}::INSTR"
    print(f"Opening {resource} ...")
    inst = rm.open_resource(resource)
    inst.timeout = timeout
    inst.read_termination = "\n"
    inst.write_termination = "\n"

    idn = inst.query("*IDN?").strip()
    print(f"  Connected: {idn}")
    return rm, inst


def list_resources(backend: str):
    """Print all detected VISA resources."""
    rm = pyvisa.ResourceManager(backend)
    resources = rm.list_resources()
    if resources:
        print("Detected VISA resources:")
        for r in resources:
            print(f"  {r}")
    else:
        print("No VISA resources found. Check adapter/driver installation.")
    rm.close()


# ---------------------------------------------------------------------------
# Data extraction from 8973A
# ---------------------------------------------------------------------------
def trigger_sweep(inst):
    """Start a single sweep and wait for completion."""
    print("  Triggering single sweep ...")
    inst.write(":INIT:CONT:ALL OFF")       # stop continuous sweep
    inst.write(":INIT:IMM")                 # trigger single sweep
    inst.write("*WAI")                       # wait for sweep to finish
    inst.query("*OPC?")                      # confirm operation complete
    print("  Sweep complete.")


def get_freq_array(inst) -> np.ndarray:
    """Build the frequency array from start/stop/points settings."""
    f_start = float(inst.query(":SENS:FREQ:STAR?").strip())
    f_stop  = float(inst.query(":SENS:FREQ:STOP?").strip())
    n_pts   = int(float(inst.query(":SENS:SWE:POIN?").strip()))
    print(f"  Freq range: {f_start/1e6:.3f} – {f_stop/1e6:.3f} MHz, {n_pts} points")
    return np.linspace(f_start, f_stop, n_pts)


def fetch_trace(inst, scpi_query: str) -> np.ndarray:
    """Fetch a comma-separated trace array from the 8973A."""
    raw = inst.query(scpi_query).strip()
    values = [float(v) for v in raw.split(",")]
    return np.array(values)


def extract_all(inst, do_trigger: bool):
    """Return freq_mhz, nf_db, gain_db arrays."""
    if do_trigger:
        trigger_sweep(inst)
    else:
        # Brief pause to make sure any in-progress sweep has settled
        time.sleep(0.5)

    freq_hz = get_freq_array(inst)
    freq_mhz = freq_hz / 1e6

    print("  Fetching corrected noise figure ...")
    nf_db = fetch_trace(inst, ":FETC:ARR:CORR:NFIG? DB")

    print("  Fetching corrected gain ...")
    gain_db = fetch_trace(inst, ":FETC:ARR:CORR:GAIN? DB")

    # Sanity check lengths
    n = len(freq_mhz)
    if len(nf_db) != n or len(gain_db) != n:
        print(f"  WARNING: length mismatch — freq={n}, NF={len(nf_db)}, "
              f"gain={len(gain_db)}.  Using shortest.")
        n = min(n, len(nf_db), len(gain_db))
        freq_mhz = freq_mhz[:n]
        nf_db = nf_db[:n]
        gain_db = gain_db[:n]

    # Replace SCPI NAN (9.91E+37) with Python NaN
    nf_db[nf_db > 9e+36] = float("nan")
    gain_db[gain_db > 9e+36] = float("nan")

    return freq_mhz, nf_db, gain_db


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def write_excel(freq_mhz, nf_db, gain_db, label, output_path, append_path):
    """Write or append data to an Excel workbook with openpyxl."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.utils import get_column_letter

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if label is None:
        label = timestamp

    # ---- decide whether to create new or append ----
    if append_path and os.path.exists(append_path):
        wb = load_workbook(append_path)
        ws = wb["Data"]
        # Find next empty column (skip col A = freq)
        next_col = ws.max_column + 1
        # Write NF
        ws.cell(row=1, column=next_col, value=f"NF (dB) – {label}")
        # Write Gain
        ws.cell(row=1, column=next_col + 1, value=f"Gain (dB) – {label}")
        for i, (nf, g) in enumerate(zip(nf_db, gain_db), start=2):
            ws.cell(row=i, column=next_col, value=round(float(nf), 4))
            ws.cell(row=i, column=next_col + 1, value=round(float(g), 4))
        out_file = append_path
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers
        ws.cell(row=1, column=1, value="Freq (MHz)")
        ws.cell(row=1, column=2, value=f"NF (dB) – {label}")
        ws.cell(row=1, column=3, value=f"Gain (dB) – {label}")

        for i, (f, nf, g) in enumerate(zip(freq_mhz, nf_db, gain_db), start=2):
            ws.cell(row=i, column=1, value=round(float(f), 4))
            ws.cell(row=i, column=2, value=round(float(nf), 4))
            ws.cell(row=i, column=3, value=round(float(g), 4))

        next_col = 4  # for chart range tracking
        out_file = output_path

    # ---- Create / update charts ----
    # Remove old chart sheets so we can rebuild
    for name in ["NF Chart", "Gain Chart"]:
        if name in wb.sheetnames:
            del wb[name]

    n_rows = len(freq_mhz) + 1  # header + data
    max_col = ws.max_column

    # Noise Figure chart
    nf_chart = ScatterChart()
    nf_chart.title = "Noise Figure vs Frequency"
    nf_chart.x_axis.title = "Frequency (MHz)"
    nf_chart.y_axis.title = "Noise Figure (dB)"
    nf_chart.width = 20
    nf_chart.height = 12
    nf_chart.style = 13

    x_ref = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    for col in range(2, max_col + 1, 2):  # NF columns are at 2, 4, 6, ...
        header = ws.cell(row=1, column=col).value or ""
        if "NF" in header:
            y_ref = Reference(ws, min_col=col, min_row=2, max_row=n_rows)
            series = Series(y_ref, x_ref, title=header)
            series.graphicalProperties.line.width = 20000  # ~1.5pt
            nf_chart.series.append(series)

    ws_nf = wb.create_sheet("NF Chart")
    ws_nf.add_chart(nf_chart, "A1")

    # Gain chart
    gain_chart = ScatterChart()
    gain_chart.title = "Gain vs Frequency"
    gain_chart.x_axis.title = "Frequency (MHz)"
    gain_chart.y_axis.title = "Gain (dB)"
    gain_chart.width = 20
    gain_chart.height = 12
    gain_chart.style = 13

    for col in range(2, max_col + 1, 2):
        header = ws.cell(row=1, column=col).value or ""
        # Gain columns are the odd ones: 3, 5, 7, ...
    for col in range(3, max_col + 1, 2):
        header = ws.cell(row=1, column=col).value or ""
        if "Gain" in header:
            y_ref = Reference(ws, min_col=col, min_row=2, max_row=n_rows)
            series = Series(y_ref, x_ref, title=header)
            series.graphicalProperties.line.width = 20000
            gain_chart.series.append(series)

    ws_gain = wb.create_sheet("Gain Chart")
    ws_gain.add_chart(gain_chart, "A1")

    # ---- Also dump a quick CSV alongside ----
    csv_path = out_file.rsplit(".", 1)[0] + ".csv"
    with open(csv_path, "w") as f:
        # Write all columns as CSV
        for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=max_col,
                                values_only=True):
            f.write(",".join(str(v) if v is not None else "" for v in row))
            f.write("\n")

    wb.save(out_file)
    print(f"\n  Excel saved:  {out_file}")
    print(f"  CSV saved:    {csv_path}")
    return out_file


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    args = parse_args()

    if args.list:
        list_resources(args.backend)
        sys.exit(0)

    # Default output filename
    if args.output is None and args.append is None:
        args.output = datetime.now().strftime("nfa_%Y%m%d_%H%M%S.xlsx")
    out_path = args.append if args.append else args.output

    rm, inst = connect(args.board, args.addr, args.timeout, args.backend)
    try:
        freq_mhz, nf_db, gain_db = extract_all(inst, args.trigger)

        print(f"\n  Points:    {len(freq_mhz)}")
        print(f"  Freq:      {freq_mhz[0]:.3f} – {freq_mhz[-1]:.3f} MHz")
        print(f"  NF range:  {np.nanmin(nf_db):.3f} – {np.nanmax(nf_db):.3f} dB")
        print(f"  Gain range: {np.nanmin(gain_db):.3f} – {np.nanmax(gain_db):.3f} dB")

        write_excel(freq_mhz, nf_db, gain_db, args.label, args.output, args.append)

    finally:
        inst.close()
        rm.close()

    print("\nDone.")


if __name__ == "__main__":
    main()
